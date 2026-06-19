from __future__ import annotations

import argparse
import datetime as dt
import html
import json
import math
import os
import re
import sqlite3
import threading
import unicodedata
import warnings
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from functools import lru_cache
from html.parser import HTMLParser
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse
from urllib.request import ProxyHandler, Request, build_opener

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parent
STATIC_DIR = ROOT_DIR / "static"
DATA_DIR = ROOT_DIR / "data"
DB_PATH = DATA_DIR / "cartas.sqlite3"
PLAY_URL = "https://playcontempladas.com.br/?segmento=todos"
UPDATE_PASSWORD = os.environ.get("SIMULADOR_UPDATE_PASSWORD") or "play123"
UPDATE_LOCK = threading.Lock()
NO_PROXY_OPENER = build_opener(ProxyHandler({}))
MARGEM_MIN = 0.95
MARGEM_MAX = 1.1
MARGEM_ENTRADA_FORMULARIO = 1.1
LIMITE_PROPOSTAS_ADM = 500
VALOR_SEM_LIMITE = 999999999.0
ADMS_SALDO_LIVRE = {"itau", "itau motos", "magalu", "porto seguro", "porto vp"}

warnings.filterwarnings("ignore", message="Unknown extension is not supported and will be removed")


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_many(values: list[object]) -> set[str]:
    return {normalized for normalized in (normalize_text(value) for value in values) if normalized}


def normalize_payload_list(value: object) -> set[str]:
    if value is None:
        return set()
    if isinstance(value, list):
        return normalize_many(value)
    return normalize_many([value])


def auto_find_workbook() -> Path:
    base = Path.home() / "OneDrive - Genec" / "Luiz Santos" / "Diversos" / "LJ Business & Automation"
    folders = sorted(base.glob("Play Cons*"))
    candidates: list[Path] = []
    for folder in folders:
        candidates.extend(folder.glob("Simulador de Cartas_v*.xlsm"))
        candidates.extend(folder.glob("Simulador de Cartas.xlsm"))

    existing = [path for path in candidates if path.exists()]
    if not existing:
        raise FileNotFoundError(
            "Nao encontrei a planilha. Informe o caminho com --workbook \"C:\\...\\arquivo.xlsm\"."
        )

    return max(existing, key=lambda path: path.stat().st_mtime)


def parse_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(float(value)):
            return None
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    text = text.replace("R$", "").replace("%", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def parse_percent_filter(value: object) -> float | None:
    number = parse_number(value)
    if number is None:
        return None
    if number > 10:
        return number / 100
    return number


def to_cents(value: float) -> int:
    return max(0, int(round(value * 100)))


def build_budget_bound_suffixes(cards: list["Card"], attr_name: str) -> list[list[tuple[int, float]]]:
    suffixes: list[list[tuple[int, float]]] = []
    for start in range(len(cards) + 1):
        items = [(to_cents(float(getattr(card, attr_name))), card.credito) for card in cards[start:]]
        items.sort(key=lambda item: float("inf") if item[0] == 0 else item[1] / item[0], reverse=True)
        suffixes.append(items)
    return suffixes


def fractional_credit_bound(items: list[tuple[int, float]], budget_cents: int) -> float:
    if budget_cents < 0:
        return 0.0

    total_credit = 0.0
    remaining_budget = budget_cents
    for cost_cents, credit in items:
        if cost_cents <= 0:
            total_credit += credit
            continue
        if cost_cents <= remaining_budget:
            total_credit += credit
            remaining_budget -= cost_cents
            continue
        if remaining_budget > 0:
            total_credit += credit * (remaining_budget / cost_cents)
        break
    return total_credit


def money(value: float | int | None) -> str:
    if value is None:
        return "-"
    quantized = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    text = f"{quantized:,.2f}"
    text = text.replace(",", "_").replace(".", ",").replace("_", ".")
    return f"R$ {text}"


def percent(value: float | None) -> str:
    if value is None:
        return "-"
    return f"{value * 100:.2f}%".replace(".", ",")


def clean_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


class HtmlTableParser(HTMLParser):
    VOID_TAGS = {"area", "base", "br", "col", "embed", "hr", "img", "input", "link", "meta", "param", "source", "track", "wbr"}

    def __init__(self, target_id: str):
        super().__init__(convert_charrefs=True)
        self.target_id = target_id
        self.in_target = False
        self.target_depth = 0
        self.in_row = False
        self.in_cell = False
        self.current_row: list[str] = []
        self.current_cell: list[str] = []
        self.rows: list[list[str]] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attrs_dict = {name: value for name, value in attrs}
        if not self.in_target and attrs_dict.get("id") == self.target_id:
            self.in_target = True
            self.target_depth = 0

        if not self.in_target:
            return

        if tag not in self.VOID_TAGS:
            self.target_depth += 1

        if tag == "tr":
            self.in_row = True
            self.current_row = []
        elif self.in_row and tag in {"td", "th"}:
            self.in_cell = True
            self.current_cell = []
        elif self.in_cell and tag == "br":
            self.current_cell.append(" ")

    def handle_endtag(self, tag: str) -> None:
        if not self.in_target:
            return

        if tag in {"td", "th"} and self.in_cell:
            text = " ".join("".join(self.current_cell).split())
            self.current_row.append(html.unescape(text))
            self.in_cell = False
            self.current_cell = []
        elif tag == "tr" and self.in_row:
            if any(cell.strip() for cell in self.current_row):
                self.rows.append(self.current_row)
            self.in_row = False
            self.current_row = []

        if tag not in self.VOID_TAGS:
            self.target_depth -= 1
            if self.target_depth <= 0:
                self.in_target = False

    def handle_data(self, data: str) -> None:
        if self.in_target and self.in_cell:
            self.current_cell.append(data)


def fetch_play_html() -> str:
    request = Request(
        PLAY_URL,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
        },
    )
    try:
        with NO_PROXY_OPENER.open(request, timeout=45) as response:
            charset = response.headers.get_content_charset() or "utf-8"
            return response.read().decode(charset, errors="replace")
    except Exception as exc:
        detail = str(exc) or repr(exc)
        raise RuntimeError(f"Nao consegui baixar os dados da Play: {detail}") from exc


def fetch_play_cards() -> list["Card"]:
    parser = HtmlTableParser("table-excel")
    parser.feed(fetch_play_html())
    if not parser.rows:
        raise ValueError("Nao encontrei a tabela de cartas no site da Play.")

    header_index, headers = find_play_header(parser.rows)
    columns = map_play_columns(headers)

    cards: list[Card] = []
    seen_codes: set[str] = set()
    for row in parser.rows[header_index + 1 :]:
        card = card_from_play_row(row, columns)
        if not card or card.codigo in seen_codes:
            continue
        seen_codes.add(card.codigo)
        cards.append(card)

    if not cards:
        raise ValueError("A tabela da Play foi localizada, mas nenhuma carta valida foi importada.")
    return cards


def find_play_header(rows: list[list[str]]) -> tuple[int, list[str]]:
    for index, row in enumerate(rows[:30]):
        normalized = [normalize_text(cell) for cell in row]
        joined = " ".join(normalized)
        if "credito" in joined and "entrada" in joined and "administradora" in joined:
            return index, row
    raise ValueError("Nao encontrei o cabecalho da tabela da Play.")


def map_play_columns(headers: list[str]) -> dict[str, int]:
    normalized = [normalize_text(header) for header in headers]

    def find(*needles: str) -> int:
        for index, header in enumerate(normalized):
            if all(needle in header for needle in needles):
                return index
        raise ValueError(f"Coluna obrigatoria nao encontrada no site da Play: {' '.join(needles)}")

    def find_parcels_count() -> int:
        for index, header in enumerate(normalized):
            if "parcela" in header and "valor" not in header and "vlr" not in header:
                return index
        raise ValueError("Coluna obrigatoria nao encontrada no site da Play: parcelas")

    def find_installment_value() -> int:
        for index, header in enumerate(normalized):
            if "parcela" in header and ("valor" in header or "vlr" in header):
                return index
        return find("parcela")

    return {
        "codigo": find("cod"),
        "categoria": find("categoria"),
        "credito": find("credito"),
        "entrada": find("entrada"),
        "num_parcelas": find_parcels_count(),
        "valor_parcela": find_installment_value(),
        "saldo_devedor": find("saldo", "devedor"),
        "administradora": find("administradora"),
        "status": find("status"),
    }


def card_from_play_row(row: list[str], columns: dict[str, int]) -> "Card | None":
    def value(name: str) -> str:
        index = columns[name]
        return row[index].strip() if index < len(row) else ""

    code = clean_code(value("codigo"))
    if not code or normalize_text(code) in {"cod", "codigo", "cod cota"}:
        return None

    credito = parse_number(value("credito")) or 0
    entrada = parse_number(value("entrada")) or 0
    valor_parcela = parse_number(value("valor_parcela")) or 0
    saldo_devedor = parse_number(value("saldo_devedor")) or 0
    num_parcelas = int(parse_number(value("num_parcelas")) or 0)

    if credito <= 0:
        return None

    return Card(
        codigo=code,
        categoria=value("categoria"),
        credito=credito,
        entrada=entrada,
        num_parcelas=num_parcelas,
        valor_parcela=valor_parcela,
        saldo_devedor=saldo_devedor,
        administradora=value("administradora"),
        status=value("status") or "Disponivel",
    )


@dataclass(frozen=True)
class Card:
    codigo: str
    categoria: str
    credito: float
    entrada: float
    num_parcelas: int
    valor_parcela: float
    saldo_devedor: float
    administradora: str
    status: str

    @property
    def custo_percent(self) -> float | None:
        if not self.credito:
            return None
        return (self.saldo_devedor + self.entrada) / self.credito

    def to_json(self) -> dict:
        return {
            "codigo": self.codigo,
            "categoria": self.categoria,
            "credito": self.credito,
            "entrada": self.entrada,
            "numParcelas": self.num_parcelas,
            "valorParcela": self.valor_parcela,
            "saldoDevedor": self.saldo_devedor,
            "administradora": self.administradora,
            "status": self.status,
            "formatted": {
                "credito": money(self.credito),
                "entrada": money(self.entrada),
                "valorParcela": money(self.valor_parcela),
                "saldoDevedor": money(self.saldo_devedor),
            },
        }


def ensure_database() -> sqlite3.Connection:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(DB_PATH)
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS cartas (
            codigo TEXT PRIMARY KEY,
            categoria TEXT NOT NULL,
            credito REAL NOT NULL,
            entrada REAL NOT NULL,
            num_parcelas INTEGER NOT NULL,
            valor_parcela REAL NOT NULL,
            saldo_devedor REAL NOT NULL,
            administradora TEXT NOT NULL,
            status TEXT NOT NULL
        )
        """
    )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS metadata (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
        """
    )
    return connection


def save_cards_to_database(cards: list[Card], source_url: str) -> str:
    updated_at = dt.datetime.now().astimezone().isoformat(timespec="seconds")
    with ensure_database() as connection:
        connection.execute("DELETE FROM cartas")
        connection.executemany(
            """
            INSERT INTO cartas (
                codigo, categoria, credito, entrada, num_parcelas,
                valor_parcela, saldo_devedor, administradora, status
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    card.codigo,
                    card.categoria,
                    card.credito,
                    card.entrada,
                    card.num_parcelas,
                    card.valor_parcela,
                    card.saldo_devedor,
                    card.administradora,
                    card.status,
                )
                for card in cards
            ],
        )
        connection.executemany(
            "INSERT OR REPLACE INTO metadata (key, value) VALUES (?, ?)",
            [
                ("source_url", source_url),
                ("updated_at", updated_at),
                ("total_cards", str(len(cards))),
            ],
        )
    return updated_at


def read_database_metadata(connection: sqlite3.Connection) -> dict[str, str]:
    rows = connection.execute("SELECT key, value FROM metadata").fetchall()
    return {str(key): str(value) for key, value in rows}


def format_updated_at(value: str | None) -> str | None:
    if not value:
        return None
    try:
        parsed = dt.datetime.fromisoformat(value)
    except ValueError:
        return value
    return parsed.strftime("%d/%m/%Y %H:%M")


class WorkbookStore:
    def __init__(self, workbook_path: Path | None):
        self.workbook_path = workbook_path
        self.cards: list[Card] = []
        self.source_label = "Sem base carregada"
        self.source_url: str | None = None
        self.updated_at: str | None = None

    def load(self) -> None:
        if DB_PATH.exists():
            self.load_from_database()
            return

        if self.workbook_path and self.workbook_path.exists():
            self.load_from_workbook()
            return

        try:
            self.update_from_play()
            return
        except Exception as exc:
            self.cards = []
            self.source_label = f"Sem base carregada: {exc}"
            self.source_url = PLAY_URL
            self.updated_at = None
            return

    def load_from_workbook(self) -> None:
        if not self.workbook_path:
            raise FileNotFoundError("Nenhuma planilha foi informada.")

        workbook = load_workbook(self.workbook_path, data_only=True, read_only=True)
        if "Dados" not in workbook.sheetnames:
            raise ValueError("A planilha precisa ter uma aba chamada 'Dados'.")

        sheet = workbook["Dados"]
        header_row, headers = self._find_header(sheet)
        columns = self._map_columns(headers)

        cards: list[Card] = []
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            code = clean_code(self._value(row, columns["codigo"]))
            if not code:
                continue

            credito = parse_number(self._value(row, columns["credito"])) or 0
            entrada = parse_number(self._value(row, columns["entrada"])) or 0
            num_parcelas = int(parse_number(self._value(row, columns["num_parcelas"])) or 0)
            valor_parcela = parse_number(self._value(row, columns["valor_parcela"])) or 0
            saldo_devedor = parse_number(self._value(row, columns["saldo_devedor"])) or 0

            cards.append(
                Card(
                    codigo=code,
                    categoria=str(self._value(row, columns["categoria"]) or "").strip(),
                    credito=credito,
                    entrada=entrada,
                    num_parcelas=num_parcelas,
                    valor_parcela=valor_parcela,
                    saldo_devedor=saldo_devedor,
                    administradora=str(self._value(row, columns["administradora"]) or "").strip(),
                    status=str(self._value(row, columns["status"]) or "").strip(),
                )
            )

        self.cards = cards
        self.source_label = f"Planilha Excel: {self.workbook_path}"
        self.source_url = None
        self.updated_at = None

    def load_from_database(self) -> None:
        with ensure_database() as connection:
            connection.row_factory = sqlite3.Row
            rows = connection.execute(
                """
                SELECT
                    codigo, categoria, credito, entrada, num_parcelas,
                    valor_parcela, saldo_devedor, administradora, status
                FROM cartas
                ORDER BY CAST(codigo AS INTEGER), codigo
                """
            ).fetchall()
            metadata = read_database_metadata(connection)

        self.cards = [
            Card(
                codigo=str(row["codigo"]),
                categoria=str(row["categoria"]),
                credito=float(row["credito"]),
                entrada=float(row["entrada"]),
                num_parcelas=int(row["num_parcelas"]),
                valor_parcela=float(row["valor_parcela"]),
                saldo_devedor=float(row["saldo_devedor"]),
                administradora=str(row["administradora"]),
                status=str(row["status"]),
            )
            for row in rows
        ]
        self.source_label = "Banco local da Play"
        self.source_url = metadata.get("source_url")
        self.updated_at = format_updated_at(metadata.get("updated_at"))

    def update_from_play(self) -> dict:
        cards = fetch_play_cards()
        save_cards_to_database(cards, PLAY_URL)
        self.load_from_database()
        return self.metadata()

    def _find_header(self, sheet) -> tuple[int, list[object]]:
        for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
            normalized = [normalize_text(value) for value in row]
            if "codigo" in normalized and any(value == "credito" for value in normalized):
                return row_index, list(row)
        raise ValueError("Nao encontrei a linha de cabecalho da aba Dados.")

    def _map_columns(self, headers: list[object]) -> dict[str, int]:
        normalized = [normalize_text(header) for header in headers]

        def find(*needles: str) -> int:
            for index, header in enumerate(normalized):
                if all(needle in header for needle in needles):
                    return index
            raise ValueError(f"Coluna obrigatoria nao encontrada: {' '.join(needles)}")

        return {
            "codigo": find("codigo"),
            "categoria": find("categoria"),
            "credito": find("credito"),
            "entrada": find("entrada"),
            "num_parcelas": find("parcelas"),
            "valor_parcela": find("valor", "parcela"),
            "saldo_devedor": find("saldo", "devedor"),
            "administradora": find("administradora"),
            "status": find("status"),
        }

    @staticmethod
    def _value(row: tuple[object, ...], index: int) -> object:
        if index >= len(row):
            return None
        return row[index]

    def filter_cards(self, params: dict[str, list[str]]) -> list[Card]:
        credito_min = parse_number(first(params, "credito_min"))
        credito_max = parse_number(first(params, "credito_max"))
        entrada_min = parse_number(first(params, "entrada_min"))
        entrada_max = parse_number(first(params, "entrada_max"))
        parcela_min = parse_number(first(params, "parcela_min"))
        parcela_max = parse_number(first(params, "parcela_max"))
        only_available = first(params, "disponivel", "1") == "1"
        query = normalize_text(first(params, "q", ""))
        categoria = normalize_text(first(params, "categoria", ""))
        administradoras = normalize_many(params.get("administradora", []))

        result: list[Card] = []
        for card in self.cards:
            if only_available and normalize_text(card.status) != "disponivel":
                continue
            if query and query not in normalize_text(
                f"{card.codigo} {card.categoria} {card.administradora} {card.status}"
            ):
                continue
            if categoria and categoria != normalize_text(card.categoria):
                continue
            if administradoras and normalize_text(card.administradora) not in administradoras:
                continue
            if credito_min is not None and card.credito < credito_min:
                continue
            if credito_max is not None and card.credito > credito_max:
                continue
            if entrada_min is not None and card.entrada < entrada_min:
                continue
            if entrada_max is not None and card.entrada > entrada_max:
                continue
            if parcela_min is not None and card.valor_parcela < parcela_min:
                continue
            if parcela_max is not None and card.valor_parcela > parcela_max:
                continue
            result.append(card)

        return result

    def calculate_proposals(self, payload: dict) -> dict:
        credito_alvo = parse_number(payload.get("creditoAlvo"))
        categoria = normalize_text(payload.get("categoria"))
        administradoras = normalize_payload_list(payload.get("administradoras"))
        saldo_regra = str(payload.get("saldo") or "INDIFERENTE").strip().upper()
        entrada_usuario = parse_number(payload.get("entradaMax"))
        parcela_max = parse_number(payload.get("parcelaMax"))
        limit = max(0, min(int(parse_number(payload.get("limit")) or 2000), 20000))

        if credito_alvo is None or credito_alvo <= 0:
            raise ValueError("Informe um credito alvo valido.")
        if not categoria:
            raise ValueError("Informe a categoria.")

        credito_min = credito_alvo * MARGEM_MIN
        credito_max = credito_alvo * MARGEM_MAX

        entrada_max = VALOR_SEM_LIMITE
        if entrada_usuario is not None:
            entrada_max = entrada_usuario * MARGEM_ENTRADA_FORMULARIO
        if parcela_max is None:
            parcela_max = VALOR_SEM_LIMITE
        entrada_limited = entrada_max < VALOR_SEM_LIMITE
        parcela_limited = parcela_max < VALOR_SEM_LIMITE

        rows: list[dict] = []
        total_propostas = 0
        total_linhas = 0
        numero_proposta = 1

        def next_id() -> str:
            nonlocal numero_proposta
            proposal_id = f"PRP-{numero_proposta:04d}"
            numero_proposta += 1
            return proposal_id

        def add_proposal(cards: list[Card]) -> None:
            nonlocal total_propostas, total_linhas
            if not cards:
                return
            proposal_id = next_id()
            total_propostas += 1
            total_linhas += len(cards)
            if len(rows) + len(cards) <= limit:
                rows.extend(build_proposal_record(proposal_id, cards)["rows"])

        eligible = [
            card
            for card in self.cards
            if self._card_base_ok(card, categoria, administradoras, saldo_regra)
        ]

        for card in eligible:
            if (
                card.credito >= credito_min
                and card.credito <= credito_max
                and card.entrada <= entrada_max
                and card.valor_parcela <= parcela_max
            ):
                add_proposal([card])

        by_admin: dict[str, list[Card]] = {}
        for card in eligible:
            if card.credito > credito_max or card.entrada > entrada_max or card.valor_parcela > parcela_max:
                continue
            key = normalize_text(card.administradora)
            if key:
                by_admin.setdefault(key, []).append(card)

        for cards in by_admin.values():
            if len(cards) < 2:
                continue
            if entrada_limited or parcela_limited:
                cards = sorted(
                    cards,
                    key=lambda card: (
                        card.entrada / card.credito if entrada_limited and card.credito else 0,
                        card.valor_parcela / card.credito if parcela_limited and card.credito else 0,
                        -card.credito,
                    ),
                )

            suffix_credito = [0.0] * (len(cards) + 1)
            for idx in range(len(cards) - 1, -1, -1):
                suffix_credito[idx] = suffix_credito[idx + 1] + cards[idx].credito

            if suffix_credito[0] < credito_min:
                continue

            entrada_bound_suffixes = build_budget_bound_suffixes(cards, "entrada") if entrada_limited else []
            parcela_bound_suffixes = build_budget_bound_suffixes(cards, "valor_parcela") if parcela_limited else []

            @lru_cache(maxsize=None)
            def max_credito_adicional(start: int, entrada_restante_centavos: int, parcela_restante_centavos: int) -> float:
                if start >= len(cards):
                    return 0.0

                limite_credito = suffix_credito[start]
                if entrada_limited:
                    limite_credito = min(
                        limite_credito,
                        fractional_credit_bound(entrada_bound_suffixes[start], entrada_restante_centavos),
                    )
                if parcela_limited:
                    limite_credito = min(
                        limite_credito,
                        fractional_credit_bound(parcela_bound_suffixes[start], parcela_restante_centavos),
                    )
                return limite_credito

            def pode_alcancar_credito(start: int, soma_credito: float, soma_entrada: float, soma_parcela: float) -> bool:
                entrada_restante = to_cents(entrada_max - soma_entrada) if entrada_limited else 0
                parcela_restante = to_cents(parcela_max - soma_parcela) if parcela_limited else 0
                return soma_credito + max_credito_adicional(start, entrada_restante, parcela_restante) >= credito_min

            generated_for_admin = 0

            def search(start: int, pacote: list[Card], soma_credito: float, soma_entrada: float, soma_parcela: float) -> None:
                nonlocal generated_for_admin
                if generated_for_admin >= LIMITE_PROPOSTAS_ADM:
                    return
                if soma_credito > credito_max:
                    return
                if soma_entrada > entrada_max:
                    return
                if soma_parcela > parcela_max:
                    return
                if not pode_alcancar_credito(start, soma_credito, soma_entrada, soma_parcela):
                    return

                if len(pacote) > 1:
                    if (
                        soma_credito >= credito_min
                        and soma_credito <= credito_max
                        and soma_entrada <= entrada_max
                        and soma_parcela <= parcela_max
                    ):
                        add_proposal(list(pacote))
                        generated_for_admin += 1

                for idx in range(start, len(cards)):
                    card = cards[idx]
                    novo_credito = soma_credito + card.credito
                    novo_entrada = soma_entrada + card.entrada
                    novo_parcela = soma_parcela + card.valor_parcela

                    if (
                        novo_credito <= credito_max
                        and novo_entrada <= entrada_max
                        and novo_parcela <= parcela_max
                    ):
                        if novo_credito < credito_min and not pode_alcancar_credito(
                            idx + 1,
                            novo_credito,
                            novo_entrada,
                            novo_parcela,
                        ):
                            continue
                        pacote.append(card)
                        search(idx + 1, pacote, novo_credito, novo_entrada, novo_parcela)
                        pacote.pop()

                    if generated_for_admin >= LIMITE_PROPOSTAS_ADM:
                        break

            search(0, [], 0, 0, 0)

        return {
            "totalPropostas": total_propostas,
            "totalLinhas": total_linhas,
            "items": rows,
            "limit": limit,
            "truncated": total_linhas > len(rows),
            "criterios": {
                "creditoAlvo": credito_alvo,
                "creditoMin": credito_min,
                "creditoMax": credito_max,
                "entradaMax": None if entrada_max == VALOR_SEM_LIMITE else entrada_max,
                "parcelaMax": None if parcela_max == VALOR_SEM_LIMITE else parcela_max,
                "saldo": saldo_regra,
                "categoria": categoria,
                "administradoras": sorted(administradoras),
            },
        }

    def _card_base_ok(self, card: Card, categoria: str, administradoras: set[str], saldo_regra: str) -> bool:
        if normalize_text(card.status) != "disponivel":
            return False
        if categoria and normalize_text(card.categoria) != categoria:
            return False
        if administradoras and normalize_text(card.administradora) not in administradoras:
            return False
        return saldo_devedor_ok(card, saldo_regra)

    def build_proposal(self, codes: list[str]) -> dict:
        code_set = {str(code).strip() for code in codes if str(code).strip()}
        selected = [card for card in self.cards if card.codigo in code_set]
        if not selected:
            raise ValueError("Nenhuma carta selecionada.")

        credito_total = sum(card.credito for card in selected)
        entrada_total = sum(card.entrada for card in selected)
        saldo_total = sum(card.saldo_devedor for card in selected)
        custo = (saldo_total + entrada_total) / credito_total if credito_total else None

        proposal = {
            "id": "WEB-0001",
            "cartas": " | ".join(card.codigo for card in selected),
            "categoria": join_unique(card.categoria for card in selected),
            "administradora": join_unique(card.administradora for card in selected),
            "parcelas": " | ".join(str(card.num_parcelas) for card in selected),
            "creditoTotal": credito_total,
            "entradaTotal": entrada_total,
            "saldoTotal": saldo_total,
            "custoPercent": custo,
            "formatted": {
                "creditoTotal": money(credito_total),
                "entradaTotal": money(entrada_total),
                "saldoTotal": money(saldo_total),
                "custoPercent": percent(custo),
            },
            "parcelasResumo": build_installment_ranges(selected),
        }
        proposal["texto"] = proposal_text(proposal)
        return proposal

    def metadata(self) -> dict:
        categorias = sorted({card.categoria for card in self.cards if card.categoria})
        administradoras = sorted({card.administradora for card in self.cards if card.administradora})
        return {
            "workbook": str(self.workbook_path or ""),
            "source": self.source_label,
            "sourceUrl": self.source_url,
            "updatedAt": self.updated_at,
            "database": str(DB_PATH) if DB_PATH.exists() else "",
            "totalCartas": len(self.cards),
            "categorias": categorias,
            "administradoras": administradoras,
        }


def build_installment_ranges(cards: list[Card]) -> list[dict]:
    max_parcels = max((card.num_parcelas for card in cards), default=0)
    if max_parcels <= 0:
        return []

    ranges: list[dict] = []
    start = 1
    current_value: float | None = None

    for month in range(1, max_parcels + 1):
        value = sum(card.valor_parcela for card in cards if card.num_parcelas >= month)
        value = round(value, 2)

        if current_value is None:
            current_value = value
            start = month
            continue

        if value != current_value:
            ranges.append(range_item(start, month - 1, current_value))
            start = month
            current_value = value

    if current_value is not None:
        ranges.append(range_item(start, max_parcels, current_value))

    return ranges


def saldo_devedor_ok(card: Card, saldo_regra: str) -> bool:
    if normalize_text(card.administradora) in ADMS_SALDO_LIVRE:
        return True
    if saldo_regra == "INDIFERENTE":
        return True
    return card.saldo_devedor <= card.credito


def card_to_proposal_json(card: Card) -> dict:
    return {
        "codigo": card.codigo,
        "categoria": card.categoria,
        "credito": card.credito,
        "entrada": card.entrada,
        "numParcelas": card.num_parcelas,
        "valorParcela": card.valor_parcela,
        "saldoDevedor": card.saldo_devedor,
        "administradora": card.administradora,
    }


def build_proposal_record(proposal_id: str, cards: list[Card]) -> dict:
    credito_total = sum(card.credito for card in cards)
    entrada_total = sum(card.entrada for card in cards)
    parcela_total = sum(card.valor_parcela for card in cards)
    saldo_total = sum(card.saldo_devedor for card in cards)
    custo = (saldo_total + entrada_total) / credito_total if credito_total else None

    rows = []
    for index, card in enumerate(cards):
        row = card_to_proposal_json(card)
        row["proposta"] = proposal_id
        if index == 0:
            row.update(
                {
                    "creditoTotal": credito_total,
                    "entradaTotal": entrada_total,
                    "parcelaTotal": parcela_total,
                    "saldoTotal": saldo_total,
                    "custoPercent": custo,
                }
            )
        rows.append(row)

    return {
        "id": proposal_id,
        "cartas": [card.codigo for card in cards],
        "rows": rows,
        "creditoTotal": credito_total,
        "entradaTotal": entrada_total,
        "parcelaTotal": parcela_total,
        "saldoTotal": saldo_total,
        "custoPercent": custo,
    }


def range_item(start: int, end: int, value: float) -> dict:
    if start == 1:
        label = f"Ate {end}"
    elif start == end:
        label = f"Parcela {start}"
    else:
        label = f"{start} a {end}"
    return {"inicio": start, "fim": end, "label": label, "valor": value, "formatted": money(value)}


def join_unique(values) -> str:
    seen: list[str] = []
    for value in values:
        text = str(value or "").strip()
        if text and text not in seen:
            seen.append(text)
    return " | ".join(seen)


def proposal_text(proposal: dict) -> str:
    lines = [
        proposal["id"],
        f"Cartas - {proposal['cartas']}",
        f"Administradora - {proposal['administradora']}",
        f"Categoria - {proposal['categoria']}",
        "",
        f"Credito Total - {proposal['formatted']['creditoTotal']}",
        f"Entrada Total - {proposal['formatted']['entradaTotal']}",
        f"Saldo Devedor Total - {proposal['formatted']['saldoTotal']}",
        f"Custo % - {proposal['formatted']['custoPercent']}",
        "",
        "Parcelas",
    ]

    for item in proposal["parcelasResumo"]:
        lines.append(f"{item['label']} - {item['formatted']}")

    return "\n".join(lines)


def first(params: dict[str, list[str]], key: str, default: str | None = None) -> str | None:
    values = params.get(key)
    if not values:
        return default
    return values[0]


class AppHandler(SimpleHTTPRequestHandler):
    store: WorkbookStore

    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(STATIC_DIR), **kwargs)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/status":
            self.send_json(self.store.metadata())
            return
        if parsed.path == "/api/cartas":
            params = parse_qs(parsed.query)
            cards = self.store.filter_cards(params)
            limit = int(first(params, "limit", "500") or "500")
            self.send_json(
                {
                    "total": len(cards),
                    "items": [card.to_json() for card in cards[:limit]],
                    "limit": limit,
                    "truncated": len(cards) > limit,
                }
            )
            return
        super().do_GET()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/reload":
            self.store.load()
            self.send_json({"ok": True, **self.store.metadata()})
            return
        if parsed.path == "/api/update-play":
            try:
                payload = self.read_json()
                senha = str(payload.get("senha", ""))
                if UPDATE_PASSWORD and senha != UPDATE_PASSWORD:
                    self.send_json({"error": "Senha invalida para atualizar os dados."}, status=HTTPStatus.UNAUTHORIZED)
                    return

                if not UPDATE_LOCK.acquire(blocking=False):
                    self.send_json({"error": "Atualizacao ja em andamento. Tente novamente em instantes."}, status=HTTPStatus.CONFLICT)
                    return

                try:
                    result = self.store.update_from_play()
                finally:
                    UPDATE_LOCK.release()

                self.send_json({"ok": True, **result})
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return
        if parsed.path == "/api/proposta":
            try:
                payload = self.read_json()
                proposal = self.store.build_proposal(payload.get("codigos", []))
                self.send_json(proposal)
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return
        if parsed.path == "/api/calcular":
            try:
                payload = self.read_json()
                result = self.store.calculate_proposals(payload)
                self.send_json(result)
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", "0"))
        raw_bytes = self.rfile.read(length)
        try:
            raw = raw_bytes.decode("utf-8")
        except UnicodeDecodeError:
            raw = raw_bytes.decode("cp1252")
        return json.loads(raw or "{}")

    def send_json(self, data: dict, status: HTTPStatus = HTTPStatus.OK) -> None:
        payload = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)


def run_server(workbook: Path | None, host: str, port: int) -> None:
    store = WorkbookStore(workbook)
    store.load()
    AppHandler.store = store
    server = ThreadingHTTPServer((host, port), AppHandler)
    print(f"Simulador web rodando em http://{host}:{port}")
    print(f"Fonte: {store.source_label}")
    print(f"Cartas carregadas: {len(store.cards)}")
    server.serve_forever()


def main() -> None:
    parser = argparse.ArgumentParser(description="MVP web do Simulador de Cartas.")
    parser.add_argument("--workbook", type=Path, default=None, help="Caminho do arquivo .xlsm/.xlsx")
    parser.add_argument("--host", default=os.environ.get("HOST", "127.0.0.1"))
    parser.add_argument("--port", type=int, default=int(os.environ.get("PORT", "8765")))
    parser.add_argument("--check", action="store_true", help="Apenas valida a leitura da planilha e sai.")
    args = parser.parse_args()

    workbook = args.workbook
    if workbook is None:
        try:
            workbook = auto_find_workbook()
        except FileNotFoundError:
            workbook = None

    if args.check:
        store = WorkbookStore(workbook)
        store.load()
        print(json.dumps(store.metadata(), ensure_ascii=False, indent=2))
        return

    run_server(workbook, args.host, args.port)


if __name__ == "__main__":
    main()
